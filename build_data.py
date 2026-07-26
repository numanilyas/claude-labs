#!/usr/bin/env python3
"""
Generates the synthetic sample-data pack for the Claude Labs curriculum.

Everything here is fictional. "Northwind Trading Co." does not exist.
The data is deliberately imperfect: the messy P&L has real formatting
crimes in it, and the bank/GL pair genuinely does not reconcile until
you find the six reconciling items.
"""
import csv
import os
import random
from datetime import date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle)

random.seed(20260625)

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs", "files")
INV = os.path.join(OUT, "invoices")
os.makedirs(INV, exist_ok=True)

D = Decimal
money = lambda x: D(str(x)).quantize(D("0.01"))


# ---------------------------------------------------------------------------
# 1. Messy P&L  ->  messy-pl-q2-fy26.xlsx
# ---------------------------------------------------------------------------
def build_messy_pl():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    grey = PatternFill("solid", fgColor="D9D9D9")

    # --- junk header block that every ERP export seems to produce ---
    ws["A1"] = "NORTHWIND TRADING CO."
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Profit & Loss  --  Quarter 2, FY2026 (Apr - Jun)"
    ws["A3"] = "CONFIDENTIAL - INTERNAL USE ONLY"
    ws["A4"] = "Exported 2026-07-06 04:12:33 by svc_reporting"
    # row 5 deliberately blank

    # --- split header across two rows, because of course ---
    ws["A6"] = "Account"
    ws["B6"] = "Apr"
    ws["D6"] = "May"
    ws["E6"] = "Jun"
    ws["F6"] = "Q2"
    ws["A7"] = ""
    ws["B7"] = "2026"
    ws["D7"] = "2026"
    ws["E7"] = "2026"
    ws["F7"] = "Total"
    for c in "ABCDEF":
        ws[f"{c}6"].font = bold
        ws[f"{c}7"].font = bold
        ws[f"{c}6"].fill = grey
        ws[f"{c}7"].fill = grey

    # column C is left entirely empty - a phantom column

    rows = [
        ("REVENUE", None, None, None, "header"),
        ("  Product Sales - Domestic", 1284500, 1341200, 1402800, "num"),
        ("  Product Sales - Export", 412300, 388900, 455100, "num"),
        ("  Service & Maintenance", 198400, 201750, 215600, "num"),
        ("  Freight Recovery", 34200, 31800, 36900, "num"),
        ("  Sales Returns & Allowances", -48200, -52400, -61300, "paren"),
        ("Total Revenue", None, None, None, "subtotal"),
        (None, None, None, None, "blank"),
        ("COST OF GOODS SOLD", None, None, None, "header"),
        ("  Materials", 612400, 648900, 701200, "num"),
        ("  Direct Labor", 224800, 231500, 248900, "text"),   # stored as TEXT
        ("  Freight In", 58200, 61400, 72800, "num"),
        ("  Inventory Adjustment", -12400, 8900, -22100, "paren"),
        ("Total COGS", None, None, None, "subtotal"),
        (None, None, None, None, "blank"),
        ("GROSS PROFIT", None, None, None, "subtotal"),
        (None, None, None, None, "blank"),
        ("OPERATING EXPENSES", None, None, None, "header"),
        ("  Salaries & Wages", 384200, 391800, 402400, "num"),
        ("  sales commissions ", 62400, 68100, 74900, "num"),   # trailing space
        ("  Rent & Occupancy", 78000, 78000, 78000, "num"),
        ("  Marketing & Advertising", 94200, 142800, 88400, "currency"),  # "$94,200"
        ("  Professional Fees", 28400, 19200, 61800, "num"),
        ("  Software & Subscriptions", 41200, 43800, 44100, "num"),
        ("  Travel & Entertainment", 22400, 31200, 28900, "num"),
        ("  Insurance", 18600, 18600, 18600, "num"),
        ("  Depreciation", 52400, 52400, 54100, "num"),
        ("  Bad Debt Expense", 8200, 4100, 38400, "num"),
        ("  Other Operating Exp", 14800, 12200, 16400, "num"),
        ("Total Operating Expenses", None, None, None, "subtotal"),
        (None, None, None, None, "blank"),
        ("OPERATING INCOME", None, None, None, "subtotal"),
        (None, None, None, None, "blank"),
        ("  Interest Expense", -18400, -18400, -18200, "paren"),
        ("  Interest Income", 2100, 2400, 2200, "num"),
        ("  FX Gain / (Loss)", 4200, -8900, 12400, "paren"),
        ("NET INCOME BEFORE TAX", None, None, None, "subtotal"),
    ]

    # compute the subtotals so the sheet is internally consistent
    def col_sums(labels):
        out = [0, 0, 0]
        for lbl, a, m, j, _k in rows:
            if lbl in labels:
                out[0] += a
                out[1] += m
                out[2] += j
        return out

    rev_lines = ["  Product Sales - Domestic", "  Product Sales - Export",
                 "  Service & Maintenance", "  Freight Recovery",
                 "  Sales Returns & Allowances"]
    cogs_lines = ["  Materials", "  Direct Labor", "  Freight In",
                  "  Inventory Adjustment"]
    opex_lines = ["  Salaries & Wages", "  sales commissions ",
                  "  Rent & Occupancy", "  Marketing & Advertising",
                  "  Professional Fees", "  Software & Subscriptions",
                  "  Travel & Entertainment", "  Insurance", "  Depreciation",
                  "  Bad Debt Expense", "  Other Operating Exp"]
    below_lines = ["  Interest Expense", "  Interest Income", "  FX Gain / (Loss)"]

    rev = col_sums(rev_lines)
    cogs = col_sums(cogs_lines)
    opex = col_sums(opex_lines)
    below = col_sums(below_lines)
    gp = [rev[i] - cogs[i] for i in range(3)]
    oi = [gp[i] - opex[i] for i in range(3)]
    nibt = [oi[i] + below[i] for i in range(3)]

    subtotals = {
        "Total Revenue": rev, "Total COGS": cogs, "GROSS PROFIT": gp,
        "Total Operating Expenses": opex, "OPERATING INCOME": oi,
        "NET INCOME BEFORE TAX": nibt,
    }

    r = 8
    for label, apr, may, jun, kind in rows:
        if kind == "blank":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label)
        if kind == "header":
            ws.cell(row=r, column=1).font = bold
            r += 1
            continue
        if kind == "subtotal":
            apr, may, jun = subtotals[label]
            ws.cell(row=r, column=1).font = bold

        vals = [apr, may, jun]
        for idx, col in enumerate([2, 4, 5]):     # B, D, E  (C stays empty)
            v = vals[idx]
            if kind == "paren" and v < 0:
                ws.cell(row=r, column=col, value=f"({abs(v):,})")
            elif kind == "text":
                ws.cell(row=r, column=col, value=f"{v}.00 ")
            elif kind == "currency":
                ws.cell(row=r, column=col, value=f"${v:,}")
            elif kind == "subtotal" and v < 0:
                ws.cell(row=r, column=col, value=f"({abs(v):,})")
            else:
                ws.cell(row=r, column=col, value=v)
        # Q2 total column, as a hardcoded value not a formula
        tot = sum(vals)
        ws.cell(row=r, column=6,
                value=f"({abs(tot):,})" if tot < 0 else tot)
        if kind == "subtotal":
            for col in [2, 4, 5, 6]:
                ws.cell(row=r, column=col).font = bold
        r += 1

    # a stray duplicated subtotal row someone pasted in and forgot
    r += 1
    ws.cell(row=r, column=1, value="Total Operating Expenses")
    ws.cell(row=r, column=2, value=opex[0])
    ws.cell(row=r, column=4, value=opex[1])
    ws.cell(row=r, column=5, value=opex[2])

    # footer junk
    r += 3
    ws.cell(row=r, column=1, value="Report generated by Northwind Financial Systems v4.2")
    ws.cell(row=r + 1, column=1, value="Questions? contact finance-reporting@northwind.example")
    ws.cell(row=r + 2, column=1,
            value="NOTE: Jun professional fees include one-time legal settlement, see memo FY26-114")

    ws.column_dimensions["A"].width = 38
    for c in "BDEF":
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["C"].width = 3

    path = os.path.join(OUT, "messy-pl-q2-fy26.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 2 + 3. Bank statement PDF  and  General ledger CSV  (they do NOT tie)
# ---------------------------------------------------------------------------
VENDORS = [
    "PACIFIC MATERIALS SUPPLY", "CASCADE LOGISTICS", "ORION SOFTWARE LTD",
    "HARBOR POINT INSURANCE", "MERIDIAN STAFFING", "BLUE RIDGE UTILITIES",
    "SUMMIT OFFICE SUPPLY", "DELTA FREIGHT SERVICES", "KESTREL LEGAL LLP",
    "NORTHGATE PROPERTY MGMT",
]
CUSTOMERS = [
    "ARGENT RETAIL GROUP", "BRIGHTWATER INDUSTRIES", "COPPERFIELD & SONS",
    "DUNMORE WHOLESALE", "EASTVALE DISTRIBUTION", "FAIRLIGHT MERCANTILE",
]

OPENING = money(742800.00)


def build_ledger_and_bank():
    shared = []          # appears in both GL and bank
    chk = 10470

    def d(day):
        return date(2026, 6, day)

    # --- receipts ---
    receipts = [
        (2, "ARGENT RETAIL GROUP", 48200.00, "INV-4471"),
        (4, "DUNMORE WHOLESALE", 22450.00, "INV-4468"),
        (8, "BRIGHTWATER INDUSTRIES", 61800.00, "INV-4475"),
        (11, "EASTVALE DISTRIBUTION", 15900.00, "INV-4479"),
        (15, "COPPERFIELD & SONS", 38400.00, "INV-4482"),
        (18, "ARGENT RETAIL GROUP", 27650.00, "INV-4488"),
        (22, "FAIRLIGHT MERCANTILE", 44100.00, "INV-4491"),
        (25, "BRIGHTWATER INDUSTRIES", 33200.00, "INV-4495"),
        (26, "DUNMORE WHOLESALE", 19800.00, "INV-4497"),
    ]
    for day, cust, amt, ref in receipts:
        shared.append({
            "date": d(day), "type": "Deposit",
            "desc": f"CUSTOMER DEPOSIT - {cust}",
            "ref": ref, "amount": money(amt),
        })

    # --- disbursements ---
    # Check numbers are assigned in date order. The 06/13 Summit Office Supply
    # payment is pulled back out below and split into two differing versions -
    # that is reconciling item (6), the transposition.
    payments = [
        (3, "NORTHGATE PROPERTY MGMT", 26000.00, "June rent"),
        (5, "PACIFIC MATERIALS SUPPLY", 84200.00, "PO-8821"),
        (5, "SUMMIT OFFICE SUPPLY", 3480.00, "PO-8834"),
        (9, "MERIDIAN STAFFING", 41900.00, "Week 22 contract labor"),
        (10, "CASCADE LOGISTICS", 18740.00, "Freight - May"),
        (12, "ORION SOFTWARE LTD", 14200.00, "Annual license"),
        (12, "BLUE RIDGE UTILITIES", 6820.00, "Utilities - May"),
        (13, "SUMMIT OFFICE SUPPLY", 1542.00, "PO-8858"),      # -> transposition
        (16, "PACIFIC MATERIALS SUPPLY", 67300.00, "PO-8849"),
        (17, "HARBOR POINT INSURANCE", 18600.00, "Q2 premium"),
        (19, "DELTA FREIGHT SERVICES", 12480.00, "Freight - Jun wk1-2"),
        (23, "MERIDIAN STAFFING", 38200.00, "Week 24 contract labor"),
        (24, "SUMMIT OFFICE SUPPLY", 2140.00, "PO-8867"),
        (24, "CASCADE LOGISTICS", 21600.00, "Freight - Jun"),
    ]
    transposed_chk = None
    for day, vend, amt, memo in sorted(payments, key=lambda p: p[0]):
        chk += 1
        if day == 13 and vend == "SUMMIT OFFICE SUPPLY":
            transposed_chk = chk
        shared.append({
            "date": d(day), "type": "Check", "desc": f"CHECK #{chk} - {vend}",
            "ref": f"CHK{chk}", "amount": -money(amt), "memo": memo,
        })
    next_chk = chk + 1

    # payroll + card, as ACH/EFT
    for day, desc, amt in [(15, "PAYROLL ACH - PERIOD ENDING 06/13", 128400.00),
                           (30, "PAYROLL ACH - PERIOD ENDING 06/27", 131200.00),
                           (20, "CORPORATE CARD SETTLEMENT", 24380.00)]:
        shared.append({"date": d(day), "type": "EFT", "desc": desc,
                       "ref": "ACH", "amount": -money(amt)})

    # =========================================================
    # The six reconciling items
    # =========================================================

    # (1)(2)(3) BANK ONLY - never made it into the GL
    bank_only = [
        {"date": d(30), "type": "Fee", "desc": "MONTHLY SERVICE CHARGE",
         "ref": "SVC", "amount": -money(85.00)},
        {"date": d(30), "type": "Interest", "desc": "INTEREST CREDIT",
         "ref": "INT", "amount": money(142.30)},
        {"date": d(21), "type": "Return", "desc":
         "RETURNED ITEM - NSF - EASTVALE DISTRIBUTION",
         "ref": "NSF", "amount": -money(3200.00)},
    ]

    # (4) GL ONLY - outstanding checks written but not yet cleared
    oc1, oc2 = next_chk, next_chk + 1
    outstanding = [
        {"date": d(27), "type": "Check", "desc": f"CHECK #{oc1} - KESTREL LEGAL LLP",
         "ref": f"CHK{oc1}", "amount": -money(4750.00), "memo": "Matter FY26-114"},
        {"date": d(29), "type": "Check",
         "desc": f"CHECK #{oc2} - DELTA FREIGHT SERVICES",
         "ref": f"CHK{oc2}", "amount": -money(1890.00), "memo": "Freight - Jun wk3-4"},
    ]

    # (5) GL ONLY - deposit in transit, hits the bank 2 July
    in_transit = [
        {"date": d(30), "type": "Deposit",
         "desc": "CUSTOMER DEPOSIT - FAIRLIGHT MERCANTILE",
         "ref": "INV-4502", "amount": money(12400.00)},
    ]

    # (6) transposition: the bank cleared 1,542.00, the GL was keyed as 1,452.00
    tx = [t for t in shared if t["ref"] == f"CHK{transposed_chk}"][0]
    shared.remove(tx)
    transposed_bank = dict(tx)                     # 1,542.00 - what actually cleared
    transposed_gl = dict(tx)
    transposed_gl["amount"] = -money(1452.00)      # what someone typed

    # (7) duplicate posting in the GL only - same JE entered twice
    dup_source = [t for t in shared if "ORION SOFTWARE" in t["desc"]][0]
    duplicate = dict(dup_source)
    duplicate["memo"] = (duplicate.get("memo", "") + " (re-posted)").strip()

    # ---- assemble ----
    gl_txns = shared + outstanding + in_transit + [transposed_gl, duplicate]
    bank_txns = shared + bank_only + [transposed_bank]

    gl_txns.sort(key=lambda t: (t["date"], t["ref"]))
    bank_txns.sort(key=lambda t: (t["date"], t["ref"]))

    gl_close = OPENING + sum(t["amount"] for t in gl_txns)
    bank_close = OPENING + sum(t["amount"] for t in bank_txns)

    # ---- prove the reconciliation actually works ----
    adj_bank = (bank_close
                + money(12400.00)                       # deposit in transit
                - money(4750.00) - money(1890.00))      # outstanding checks
    adj_book = (gl_close
                + money(142.30)                         # interest not booked
                - money(85.00)                          # service charge
                - money(3200.00)                        # NSF
                + money(dup_source["amount"]) * -1      # reverse the duplicate
                - money(90.00))                         # transposition 1452 -> 1542
    assert adj_bank == adj_book, f"recon broken: {adj_bank} vs {adj_book}"

    # ---- write GL csv ----
    gl_path = os.path.join(OUT, "general-ledger-june-2026.csv")
    with open(gl_path, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Entry No", "Date", "Account", "Account Name",
                    "Description", "Reference", "Debit", "Credit", "Memo"])
        bal = OPENING
        for i, t in enumerate(gl_txns, start=1):
            amt = t["amount"]
            bal += amt
            w.writerow([
                f"JE-{2600 + i}",
                t["date"].isoformat(),
                "1010", "Cash - Operating",
                t["desc"],
                t["ref"],
                f"{amt:.2f}" if amt > 0 else "",
                f"{abs(amt):.2f}" if amt < 0 else "",
                t.get("memo", ""),
            ])

    # ---- write bank statement pdf ----
    bank_path = os.path.join(OUT, "bank-statement-june-2026.pdf")
    _bank_pdf(bank_path, bank_txns, bank_close)

    return gl_path, bank_path, gl_close, bank_close, adj_bank


def _bank_pdf(path, txns, closing):
    doc = SimpleDocTemplate(path, pagesize=LETTER,
                            leftMargin=0.6 * inch, rightMargin=0.6 * inch,
                            topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                            title="Statement of Account - June 2026")
    ss = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=ss["Heading1"], fontSize=16,
                       textColor=colors.HexColor("#1a3a5c"), spaceAfter=2)
    small = ParagraphStyle("s", parent=ss["Normal"], fontSize=8,
                           textColor=colors.HexColor("#555555"))
    story = []

    story.append(Paragraph("MERIDIAN COMMERCE BANK", h))
    story.append(Paragraph("Statement of Account", ss["Heading3"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "NORTHWIND TRADING CO.<br/>2200 Cannery Row, Suite 400<br/>"
        "Portland, OR 97209", small))
    story.append(Spacer(1, 8))

    debits = sum(t["amount"] for t in txns if t["amount"] < 0)
    credits = sum(t["amount"] for t in txns if t["amount"] > 0)

    summary = [
        ["Account Number", "****-4471-0092"],
        ["Statement Period", "June 1, 2026 - June 30, 2026"],
        ["Opening Balance", f"{OPENING:,.2f}"],
        ["Total Credits", f"{credits:,.2f}"],
        ["Total Debits", f"({abs(debits):,.2f})"],
        ["Closing Balance", f"{closing:,.2f}"],
    ]
    t = Table(summary, colWidths=[1.9 * inch, 2.1 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, -2), (-1, -2), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    data = [["Date", "Description", "Reference", "Debit", "Credit", "Balance"]]
    bal = OPENING
    for tx in txns:
        bal += tx["amount"]
        data.append([
            tx["date"].strftime("%m/%d"),
            tx["desc"],
            tx["ref"],
            f"{abs(tx['amount']):,.2f}" if tx["amount"] < 0 else "",
            f"{tx['amount']:,.2f}" if tx["amount"] > 0 else "",
            f"{bal:,.2f}",
        ])

    tbl = Table(data, colWidths=[0.55 * inch, 3.0 * inch, 0.85 * inch,
                                 0.9 * inch, 0.9 * inch, 1.0 * inch],
                repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f2f5f8")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "Please examine this statement and report any discrepancy within 30 days. "
        "Meridian Commerce Bank, Member FDIC. This is a fictional statement "
        "created for training purposes.", small))
    doc.build(story)


# ---------------------------------------------------------------------------
# 4. Budget vs actual
# ---------------------------------------------------------------------------
def build_budget_actual():
    depts = ["Sales", "Marketing", "Operations", "Engineering", "G&A",
             "Customer Success"]
    accounts = {
        "Salaries & Wages": (180000, 0.02),
        "Contract Labor": (28000, 0.25),
        "Travel & Entertainment": (12000, 0.35),
        "Software & Subscriptions": (16000, 0.12),
        "Marketing Programs": (40000, 0.30),
        "Professional Fees": (9000, 0.45),
        "Facilities Allocation": (14000, 0.03),
        "Other": (6000, 0.30),
    }
    months = ["2026-04", "2026-05", "2026-06"]
    scale = {"Sales": 1.0, "Marketing": 0.55, "Operations": 1.35,
             "Engineering": 1.6, "G&A": 0.7, "Customer Success": 0.65}

    # planted stories the labs ask people to find
    spikes = {
        ("Engineering", "Contract Labor", "2026-06"): 3.1,
        ("Marketing", "Marketing Programs", "2026-05"): 1.9,
        ("G&A", "Professional Fees", "2026-06"): 4.4,
        ("Operations", "Travel & Entertainment", "2026-06"): 0.35,
        ("Sales", "Contract Labor", "2026-04"): 0.28,
    }

    path = os.path.join(OUT, "budget-vs-actual-q2-fy26.csv")
    with open(path, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Month", "Department", "Account", "Budget", "Actual"])
        for m in months:
            for dept in depts:
                for acct, (base, vol) in accounts.items():
                    if acct == "Marketing Programs" and dept != "Marketing":
                        continue
                    budget = round(base * scale[dept] / 3, -2)
                    if budget <= 0:
                        continue
                    mult = spikes.get((dept, acct, m))
                    if mult is None:
                        mult = 1 + random.uniform(-vol, vol) * 0.45
                    actual = round(budget * mult, 2)
                    w.writerow([m, dept, acct, f"{budget:.2f}", f"{actual:.2f}"])
    return path


# ---------------------------------------------------------------------------
# 5. AR aging
# ---------------------------------------------------------------------------
def build_ar_aging():
    path = os.path.join(OUT, "ar-aging-june-2026.csv")
    asof = date(2026, 6, 30)
    rows = []
    inv = 4380
    for cust in CUSTOMERS + ["GRANITE BAY OUTFITTERS", "HOLLOWAY SUPPLY CO",
                             "IRONWOOD PARTNERS", "JUNIPER MERCANTILE"]:
        for _ in range(random.randint(2, 6)):
            inv += random.randint(1, 4)
            days = random.choices([12, 25, 40, 55, 75, 100, 140],
                                  weights=[30, 25, 15, 12, 8, 6, 4])[0]
            issued = asof - timedelta(days=days)
            amt = round(random.uniform(1800, 62000), 2)
            bucket = ("Current" if days <= 30 else
                      "31-60" if days <= 60 else
                      "61-90" if days <= 90 else "90+")
            rows.append([f"INV-{inv}", cust, issued.isoformat(),
                         (issued + timedelta(days=30)).isoformat(),
                         f"{amt:.2f}", days - 30 if days > 30 else 0, bucket])
    rows.sort(key=lambda r: (r[1], r[2]))
    with open(path, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Invoice", "Customer", "Invoice Date", "Due Date",
                    "Amount", "Days Past Due", "Aging Bucket"])
        w.writerows(rows)
    return path


# ---------------------------------------------------------------------------
# 6. Vendor invoices - five PDFs, five different layouts
# ---------------------------------------------------------------------------
INVOICE_SPECS = [
    dict(file="invoice-pacific-materials.pdf", vendor="Pacific Materials Supply",
         addr="1180 Industrial Way<br/>Tacoma, WA 98421",
         num="PMS-2026-3391", date="2026-06-16", terms="Net 30",
         due="2026-07-16", po="PO-8849", accent="#0b5d3b",
         lines=[("Cold-rolled steel sheet, 16ga", "1,200 sheet", 42.50),
                ("Aluminum extrusion, 6061-T6", "480 m", 28.75),
                ("Fasteners, assorted", "1 lot", 3120.00),
                ("Freight surcharge", "1", 1840.00)],
         tax=0.0),
    dict(file="invoice-cascade-logistics.pdf", vendor="Cascade Logistics",
         addr="88 Terminal Road<br/>Portland, OR 97218",
         num="CL-77412", date="2026-06-24", terms="Net 15",
         due="2026-07-09", po="", accent="#8a4b0f",
         lines=[("LTL freight - Zone 4 (14 shipments)", "14", 890.00),
                ("Fuel surcharge", "1", 2840.00),
                ("Detention charges", "6 hr", 145.00),
                ("Liftgate service", "9", 75.00)],
         tax=0.0),
    dict(file="invoice-orion-software.pdf", vendor="Orion Software Ltd",
         addr="Unit 12, Kestrel House<br/>Reading RG1 8QT, United Kingdom",
         num="ORN-INV-20260612", date="2026-06-12", terms="Due on receipt",
         due="2026-06-12", po="PO-8855", accent="#2b3a8f",
         lines=[("Orion Analytics Platform - annual, 40 seats", "40", 310.00),
                ("Premium support tier", "1", 1800.00)],
         tax=0.20),
    dict(file="invoice-kestrel-legal.pdf", vendor="Kestrel Legal LLP",
         addr="One Court Square, 22nd Floor<br/>Seattle, WA 98104",
         num="KL-9920-06", date="2026-06-27", terms="Net 30",
         due="2026-07-27", po="", accent="#4a1f47",
         lines=[("Partner time - matter FY26-114", "18.5 hr", 685.00),
                ("Associate time - matter FY26-114", "42.0 hr", 395.00),
                ("Paralegal time", "16.0 hr", 165.00),
                ("Filing fees and disbursements", "1", 2140.00)],
         tax=0.0),
    dict(file="invoice-meridian-staffing.pdf", vendor="Meridian Staffing",
         addr="4400 Commerce Center Dr<br/>Beaverton, OR 97006",
         num="MS-2026-06-B", date="2026-06-23", terms="Net 7",
         due="2026-06-30", po="PO-8861", accent="#1a3a5c",
         lines=[("Contract engineer - week 24 (3 resources)", "120 hr", 118.00),
                ("Contract QA analyst - week 24", "40 hr", 86.00),
                ("Warehouse temp labor - week 24", "168 hr", 34.50),
                ("Agency placement fee", "1", 4200.00)],
         tax=0.0),
]


def build_invoices():
    ss = getSampleStyleSheet()
    paths = []
    for spec in INVOICE_SPECS:
        p = os.path.join(INV, spec["file"])
        accent = colors.HexColor(spec["accent"])
        doc = SimpleDocTemplate(p, pagesize=LETTER,
                                leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                                topMargin=0.7 * inch, bottomMargin=0.7 * inch,
                                title=f"Invoice {spec['num']}")
        h = ParagraphStyle("h", parent=ss["Heading1"], fontSize=17,
                           textColor=accent, spaceAfter=1)
        small = ParagraphStyle("s", parent=ss["Normal"], fontSize=8.5, leading=11)
        right = ParagraphStyle("r", parent=small, alignment=2)

        story = [Paragraph(spec["vendor"].upper(), h),
                 Paragraph(spec["addr"], small), Spacer(1, 14)]

        meta = [["INVOICE", spec["num"]],
                ["Invoice Date", spec["date"]],
                ["Terms", spec["terms"]],
                ["Due Date", spec["due"]]]
        if spec["po"]:
            meta.append(["Your PO", spec["po"]])
        meta.append(["Bill To", "Northwind Trading Co."])
        mt = Table(meta, colWidths=[1.3 * inch, 2.4 * inch])
        mt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TEXTCOLOR", (1, 0), (1, 0), accent),
            ("FONTNAME", (1, 0), (1, 0), "Helvetica-Bold"),
        ]))
        story += [mt, Spacer(1, 16)]

        data = [["Description", "Qty", "Unit Price", "Amount"]]
        sub = 0.0
        for desc, qty, unit in spec["lines"]:
            n = float(qty.split()[0].replace(",", "")) if qty[0].isdigit() else 1.0
            amt = n * unit
            sub += amt
            data.append([desc, qty, f"{unit:,.2f}", f"{amt:,.2f}"])
        tax = sub * spec["tax"]
        total = sub + tax
        data.append(["", "", "Subtotal", f"{sub:,.2f}"])
        if spec["tax"]:
            data.append(["", "", f"VAT @ {int(spec['tax']*100)}%", f"{tax:,.2f}"])
        data.append(["", "", "TOTAL DUE", f"{total:,.2f}"])

        n_line = len(spec["lines"])
        tbl = Table(data, colWidths=[3.6 * inch, 1.0 * inch, 1.1 * inch, 1.3 * inch])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 1), (-1, n_line), 0.25, colors.HexColor("#cccccc")),
            ("LINEABOVE", (2, n_line + 1), (-1, n_line + 1), 0.75, accent),
            ("FONTNAME", (2, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story += [tbl, Spacer(1, 18),
                  Paragraph("Remit payment to the account on file. "
                            "Late payments accrue 1.5% monthly. "
                            "<i>Fictional invoice, created for training.</i>", small)]
        doc.build(story)
        paths.append(p)
    return paths


# ---------------------------------------------------------------------------
# 7. MySQL sample database  ->  docs/files/db/northwind-setup.sql
#
# Derived from the CSVs above so every figure in the Data Track ties back to
# the figures in the Finance Track. Deliberate problems, planted on purpose:
#
#   * invoices.customer_id is NULL for INV-4391 (22,010.93). An INNER JOIN to
#     customers silently drops it, so the AR total comes out 22,010.93 light.
#   * customers.region is NULL for two customers and '' (empty string) for a
#     third, so `WHERE region IS NULL` finds two of the three.
#   * gl_entries carries the duplicate cheque #10476 posting (JE-2610 and
#     JE-2611) that the Finance Track bank rec turns up.
#   * budget_actual_raw is created empty and denormalised - the learner
#     imports the CSV into it, then normalises it themselves.
# ---------------------------------------------------------------------------

DB_ORPHAN_INVOICE = "INV-4391"          # customer_id deliberately left NULL

CUSTOMER_MASTER = [
    # id,     name,                     region,  terms, credit limit, manager
    ("C001", "Argent Retail Group",     "North",   30, 100000.00, "R. Aldridge"),
    ("C002", "Brightwater Industries",  "South",   30, 150000.00, "R. Aldridge"),
    ("C003", "Copperfield & Sons",      None,      30, 100000.00, "M. Okafor"),
    ("C004", "Dunmore Wholesale",       "East",    30, 150000.00, "M. Okafor"),
    ("C005", "Eastvale Distribution",   "East",    30, 150000.00, "M. Okafor"),
    ("C006", "Fairlight Mercantile",    "West",    30, 125000.00, "S. K. Rao"),
    ("C007", "Granite Bay Outfitters",  "",        30, 100000.00, "S. K. Rao"),
    ("C008", "Holloway Supply Co",      "North",   30, 125000.00, "R. Aldridge"),
    ("C009", "Ironwood Partners",       None,      30, 100000.00, "S. K. Rao"),
    ("C010", "Juniper Mercantile",      "South",   30, 125000.00, "M. Okafor"),
]

# CSV customer strings are upper case; map them onto the master above.
_CUST_BY_UPPER = {c[1].upper(): c[0] for c in CUSTOMER_MASTER}


def _sqlstr(v):
    """Render a Python value as a MySQL literal."""
    if v is None:
        return "NULL"
    if isinstance(v, (int, float)):
        return str(v)
    return "'" + str(v).replace("\\", "\\\\").replace("'", "''") + "'"


def build_database():
    import csv as _csv
    from decimal import Decimal

    db_dir = os.path.join(OUT, "db")
    os.makedirs(db_dir, exist_ok=True)
    path = os.path.join(db_dir, "northwind-setup.sql")

    ar = list(_csv.DictReader(open(os.path.join(OUT, "ar-aging-june-2026.csv"),
                                   encoding="utf-8")))
    gl = list(_csv.DictReader(open(os.path.join(OUT, "general-ledger-june-2026.csv"),
                                   encoding="utf-8")))

    # ---- invoices -------------------------------------------------------
    invoices = []
    for r in ar:
        cid = _CUST_BY_UPPER[r["Customer"].strip().upper()]
        if r["Invoice"] == DB_ORPHAN_INVOICE:
            cid = None                       # the planted orphan
        invoices.append((r["Invoice"], cid, r["Invoice Date"], r["Due Date"],
                         Decimal(r["Amount"]), int(r["Days Past Due"]),
                         r["Aging Bucket"]))

    total_ar = sum(i[4] for i in invoices)
    joined_ar = sum(i[4] for i in invoices if i[1] is not None)
    orphan_amt = total_ar - joined_ar

    assert len(invoices) == 34, "expected 34 invoices"
    assert total_ar == Decimal("1096352.29"), f"AR total moved: {total_ar}"
    assert orphan_amt == Decimal("22010.93"), f"orphan amount moved: {orphan_amt}"
    assert joined_ar == Decimal("1074341.36"), f"joined AR moved: {joined_ar}"

    over90 = sum(i[4] for i in invoices if i[6] == "90+")
    over90_joined = sum(i[4] for i in invoices if i[6] == "90+" and i[1] is not None)
    assert over90 == Decimal("116829.32"), f"90+ total moved: {over90}"
    assert over90_joined == Decimal("94818.39"), f"90+ joined moved: {over90_joined}"

    # ---- credit limits: exactly three customers must be over -------------
    exposure = {}
    for inv in invoices:
        if inv[1]:
            exposure[inv[1]] = exposure.get(inv[1], Decimal("0")) + inv[4]
    over_limit = sorted(c[0] for c in CUSTOMER_MASTER
                        if exposure.get(c[0], Decimal("0")) > Decimal(str(c[4])))
    assert over_limit == ["C003", "C004", "C009"], f"over-limit set moved: {over_limit}"

    # ---- region nulls ----------------------------------------------------
    n_null = sum(1 for c in CUSTOMER_MASTER if c[2] is None)
    n_blank = sum(1 for c in CUSTOMER_MASTER if c[2] == "")
    assert (n_null, n_blank) == (2, 1), "region trap moved"

    # ---- gl entries ------------------------------------------------------
    gl_rows = []
    for r in gl:
        gl_rows.append((r["Entry No"], r["Date"], r["Account"], r["Account Name"],
                        r["Description"], r["Reference"],
                        Decimal(r["Debit"]) if r["Debit"] else None,
                        Decimal(r["Credit"]) if r["Credit"] else None,
                        r["Memo"] or None))
    assert len(gl_rows) == 30, "expected 30 GL rows"

    dup_refs = {}
    for g in gl_rows:
        if g[7] is not None:
            dup_refs.setdefault(g[5], []).append(g[7])
    dups = {k: v for k, v in dup_refs.items() if len(v) > 1 and k.startswith("CHK")}
    assert list(dups) == ["CHK10476"], f"duplicate cheque trap moved: {list(dups)}"
    assert sum(dups["CHK10476"]) == Decimal("28400.00")

    gl_debits = sum(g[6] for g in gl_rows if g[6])
    gl_credits = sum(g[7] for g in gl_rows if g[7])

    # ---- write the script ------------------------------------------------
    L = []
    w = L.append
    w("-- Northwind Trading Co. - sample database for the Claude Labs Data Track")
    w("-- Fictional data. Generated by build_data.py - do not edit by hand.")
    w("-- Tested against MySQL 8.x. Run with:  SOURCE northwind-setup.sql")
    w("")
    w("DROP DATABASE IF EXISTS northwind;")
    w("CREATE DATABASE northwind;")
    w("USE northwind;")
    w("")
    w("-- ---------------------------------------------------------------")
    w("-- Customer master. One row per customer - the whole point of a")
    w("-- relational database is that this name is stored exactly once.")
    w("-- ---------------------------------------------------------------")
    w("CREATE TABLE customers (")
    w("  customer_id        CHAR(4)       NOT NULL,")
    w("  customer_name      VARCHAR(80)   NOT NULL,")
    w("  region             VARCHAR(20)       NULL,")
    w("  payment_terms_days INT           NOT NULL DEFAULT 30,")
    w("  credit_limit       DECIMAL(12,2) NOT NULL,")
    w("  account_manager    VARCHAR(40)       NULL,")
    w("  PRIMARY KEY (customer_id)")
    w(") ENGINE=InnoDB;")
    w("")
    w("INSERT INTO customers")
    w("  (customer_id, customer_name, region, payment_terms_days, credit_limit, account_manager)")
    w("VALUES")
    vals = [
        "  (%s, %s, %s, %d, %.2f, %s)" % (_sqlstr(c[0]), _sqlstr(c[1]),
                                          _sqlstr(c[2]), c[3], c[4], _sqlstr(c[5]))
        for c in CUSTOMER_MASTER
    ]
    w(",\n".join(vals) + ";")
    w("")
    w("-- ---------------------------------------------------------------")
    w("-- Open sales invoices at 30 June 2026. customer_id is a foreign key")
    w("-- back to customers - and it is NULL on one row, which is on purpose.")
    w("-- ---------------------------------------------------------------")
    w("CREATE TABLE invoices (")
    w("  invoice_no     VARCHAR(12)   NOT NULL,")
    w("  customer_id    CHAR(4)           NULL,")
    w("  invoice_date   DATE          NOT NULL,")
    w("  due_date       DATE          NOT NULL,")
    w("  amount         DECIMAL(12,2) NOT NULL,")
    w("  days_past_due  INT           NOT NULL,")
    w("  aging_bucket   VARCHAR(10)   NOT NULL,")
    w("  PRIMARY KEY (invoice_no),")
    w("  CONSTRAINT fk_invoices_customer")
    w("    FOREIGN KEY (customer_id) REFERENCES customers (customer_id)")
    w(") ENGINE=InnoDB;")
    w("")
    w("INSERT INTO invoices")
    w("  (invoice_no, customer_id, invoice_date, due_date, amount, days_past_due, aging_bucket)")
    w("VALUES")
    vals = [
        "  (%s, %s, %s, %s, %s, %d, %s)" % (_sqlstr(i[0]), _sqlstr(i[1]),
                                            _sqlstr(i[2]), _sqlstr(i[3]),
                                            f"{i[4]:.2f}", i[5], _sqlstr(i[6]))
        for i in invoices
    ]
    w(",\n".join(vals) + ";")
    w("")
    w("-- ---------------------------------------------------------------")
    w("-- June cash ledger. Note account_name is repeated on every single")
    w("-- row - that is what a denormalised table looks like.")
    w("-- ---------------------------------------------------------------")
    w("CREATE TABLE gl_entries (")
    w("  entry_no     VARCHAR(10)   NOT NULL,")
    w("  entry_date   DATE          NOT NULL,")
    w("  account_code VARCHAR(10)   NOT NULL,")
    w("  account_name VARCHAR(60)   NOT NULL,")
    w("  description  VARCHAR(120)  NOT NULL,")
    w("  reference    VARCHAR(20)       NULL,")
    w("  debit        DECIMAL(12,2)     NULL,")
    w("  credit       DECIMAL(12,2)     NULL,")
    w("  memo         VARCHAR(120)      NULL,")
    w("  PRIMARY KEY (entry_no)")
    w(") ENGINE=InnoDB;")
    w("")
    w("INSERT INTO gl_entries")
    w("  (entry_no, entry_date, account_code, account_name, description, reference, debit, credit, memo)")
    w("VALUES")
    vals = []
    for g in gl_rows:
        vals.append("  (%s, %s, %s, %s, %s, %s, %s, %s, %s)" % (
            _sqlstr(g[0]), _sqlstr(g[1]), _sqlstr(g[2]), _sqlstr(g[3]),
            _sqlstr(g[4]), _sqlstr(g[5]),
            "NULL" if g[6] is None else f"{g[6]:.2f}",
            "NULL" if g[7] is None else f"{g[7]:.2f}",
            _sqlstr(g[8])))
    w(",\n".join(vals) + ";")
    w("")
    w("-- ---------------------------------------------------------------")
    w("-- Empty on purpose. Lab 2 imports budget-vs-actual-q2-fy26.csv into")
    w("-- this table; Lab 5 splits it into proper tables.")
    w("-- ---------------------------------------------------------------")
    w("CREATE TABLE budget_actual_raw (")
    w("  month      VARCHAR(7)    NOT NULL,")
    w("  department VARCHAR(40)   NOT NULL,")
    w("  account    VARCHAR(60)   NOT NULL,")
    w("  budget     DECIMAL(12,2) NOT NULL,")
    w("  actual     DECIMAL(12,2) NOT NULL")
    w(") ENGINE=InnoDB;")
    w("")
    w("SELECT 'northwind is ready' AS status,")
    w("       (SELECT COUNT(*) FROM customers)  AS customers,")
    w("       (SELECT COUNT(*) FROM invoices)   AS invoices,")
    w("       (SELECT COUNT(*) FROM gl_entries) AS gl_entries;")
    w("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L))

    return path, {
        "total_ar": total_ar, "joined_ar": joined_ar, "orphan": orphan_amt,
        "over90": over90, "over90_joined": over90_joined,
        "gl_debits": gl_debits, "gl_credits": gl_credits,
        "over_limit": over_limit,
    }


if __name__ == "__main__":
    pl = build_messy_pl()
    gl, bank, gl_close, bank_close, adj = build_ledger_and_bank()
    ba = build_budget_actual()
    ar = build_ar_aging()
    invs = build_invoices()
    dbsql, dbstats = build_database()

    print(f"P&L            {pl}")
    print(f"GL             {gl}")
    print(f"Bank stmt      {bank}")
    print(f"Budget/actual  {ba}")
    print(f"AR aging       {ar}")
    for i in invs:
        print(f"Invoice        {i}")
    print()
    print(f"  GL closing balance    {gl_close:>14,.2f}")
    print(f"  Bank closing balance  {bank_close:>14,.2f}")
    print(f"  Difference            {gl_close - bank_close:>14,.2f}")
    print(f"  Adjusted (both sides)  {adj:>14,.2f}")
    print()
    print(f"DB script      {dbsql}")
    for k, v in dbstats.items():
        print(f"  {k:<16} {v}")
